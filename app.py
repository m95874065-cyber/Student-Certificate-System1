import streamlit as st
from datetime import datetime
import pandas as pd
import plotly.express as px
from supabase import create_client
import hashlib
import secrets
from io import BytesIO
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle


# ============================================================
# PAGE CONFIGURATION
# ============================================================

st.set_page_config(
    page_title="Student Certificate Management System",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)


# ============================================================
# SUPABASE CONFIGURATION
# ============================================================

SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]

supabase = create_client(
    SUPABASE_URL,
    SUPABASE_KEY
)

STORAGE_BUCKET = "certificates"


# ============================================================
# PASSWORD SECURITY
# ============================================================

def hash_password(password):

    salt = secrets.token_hex(16)

    password_hash = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt.encode("utf-8"),
        100000
    ).hex()

    return f"{salt}${password_hash}"


def verify_password(password, stored_password):

    try:

        salt, stored_hash = stored_password.split("$")

        password_hash = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            salt.encode("utf-8"),
            100000
        ).hex()

        return secrets.compare_digest(
            password_hash,
            stored_hash
        )

    except ValueError:

        return False


# ============================================================
# DATABASE FUNCTIONS
# ============================================================

def get_students():

    response = (
        supabase
        .table("students")
        .select("*")
        .execute()
    )

    return response.data or []


def get_student(register_no):

    response = (
        supabase
        .table("students")
        .select("*")
        .eq(
            "register_no",
            register_no
        )
        .execute()
    )

    if response.data:

        return response.data[0]

    return None


def get_linkedin_profile_url(student):

    return student.get("linkedin_profile_url") or ""


def update_linkedin_profile(register_no, linkedin_profile_url):

    (
        supabase
        .table("students")
        .update({
            "linkedin_profile_url": linkedin_profile_url
        })
        .eq(
            "register_no",
            register_no
        )
        .execute()
    )


def get_linkedin_status(certificate):

    return certificate.get("linkedin_status") or "Not Updated"


def update_linkedin_status(certificate_id, status):

    (
        supabase
        .table("certificates")
        .update({
            "linkedin_status": status
        })
        .eq(
            "id",
            certificate_id
        )
        .execute()
    )


# ============================================================
# CERTIFICATE DATABASE FUNCTIONS
# ============================================================

def get_certificates(register_no=None):

    query = (
        supabase
        .table("certificates")
        .select("*")
    )

    if register_no:

        query = query.eq(
            "register_no",
            register_no
        )

    response = query.execute()

    return response.data or []


# ============================================================
# ACHIEVEMENT BADGES
# ============================================================

def get_achievement_badge(completed, total):

    if total == 0:

        return (
            "🌱 Getting Started",
            "Start completing your certificates."
        )

    progress = completed / total

    if progress == 1:

        return (
            "🏆 Certificate Champion",
            "Amazing! All certificates completed."
        )

    elif progress >= 0.75:

        return (
            "🥇 Skill Master",
            "Excellent progress! Keep going."
        )

    elif progress >= 0.50:

        return (
            "🥈 Skill Builder",
            "Great progress! You are halfway there."
        )

    elif progress > 0:

        return (
            "🏅 Beginner Achiever",
            "Good start! Keep completing certificates."
        )

    else:

        return (
            "🌱 Getting Started",
            "Complete your first certificate to earn a badge."
        )


# ============================================================
# REGISTER NUMBER SORTING
# ============================================================

def get_numeric_register_number(register_no):

    try:

        if "BAI" in register_no:

            number_part = (
                register_no
                .split("BAI")[-1]
            )

            return int(number_part)

        return 999999999

    except:

        return 999999999


# ============================================================
# STORAGE FILE PATH
# ============================================================

def storage_file_path(
    register_no,
    certificate_name,
    original_name
):

    safe_name = (
        certificate_name
        .replace(" ", "_")
    )

    return (
        register_no
        + "_"
        + safe_name
        + "_"
        + original_name
    )


def get_uploaded_certificate_files(register_no, certificate_name):
    prefix = f"{register_no}_{certificate_name.replace(' ', '_')}_"
    try:
        files = (
            supabase.storage
            .from_(STORAGE_BUCKET)
            .list()
        ) or []
        return [
            f.get("name") for f in files
            if f.get("name", "").startswith(prefix)
            and f.get("name")
        ]
    except Exception:
        return []


def download_storage_file(file_name):
    return (
        supabase.storage
        .from_(STORAGE_BUCKET)
        .download(file_name)
    )


def make_excel_report(student, certificates):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Report"

    ws["A1"] = "Student Certificate Management System"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:F1")
    ws["A2"] = "Student Certificate Report"
    ws["A2"].font = Font(bold=True, size=13)
    ws.merge_cells("A2:F2")

    details = [
        ("Name", student.get("name", "")),
        ("Register Number", student.get("register_no", "")),
        ("Department", student.get("department", "")),
        ("Total Certificates", len(certificates)),
        ("Completed", sum(1 for c in certificates if c.get("status") == "Completed")),
        ("Pending", sum(1 for c in certificates if c.get("status") == "Pending")),
        ("Not Updated", sum(1 for c in certificates if c.get("status") == "Status")),
    ]
    row = 4
    for label, value in details:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    headers = ["S.No", "Certificate", "Status", "Deadline", "LinkedIn Status", "LinkedIn URL"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for i, certificate in enumerate(certificates, 1):
        values = [
            i,
            certificate.get("certificate_name", ""),
            "Not Updated" if certificate.get("status") == "Status" else certificate.get("status", ""),
            certificate.get("deadline", ""),
            certificate.get("linkedin_status") or "Not Updated",
            certificate.get("linkedin_url") or "",
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    widths = {1: 8, 2: 32, 3: 18, 4: 16, 5: 20, 6: 55}
    for col, width in widths.items():
        ws.column_dimensions[chr(64 + col)].width = width

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def make_pdf_report(student, certificates):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=35, leftMargin=35, topMargin=35, bottomMargin=35)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("ReportTitle", parent=styles["Title"], alignment=TA_CENTER, fontSize=18, spaceAfter=12)
    story = [
        Paragraph("Student Certificate Management System", title),
        Paragraph("Student Certificate Report", ParagraphStyle("Sub", parent=styles["Heading2"], alignment=TA_CENTER)),
        Spacer(1, 12),
    ]

    student_data = [
        ["Name", student.get("name", "")],
        ["Register Number", student.get("register_no", "")],
        ["Department", student.get("department", "")],
        ["Total Certificates", str(len(certificates))],
        ["Completed", str(sum(1 for c in certificates if c.get("status") == "Completed"))],
        ["Pending", str(sum(1 for c in certificates if c.get("status") == "Pending"))],
        ["Not Updated", str(sum(1 for c in certificates if c.get("status") == "Status"))],
    ]
    table = Table(student_data, colWidths=[140, 350])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (0,-1), colors.lightgrey),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story += [table, Spacer(1, 15)]

    data = [["S.No", "Certificate", "Status", "Deadline", "LinkedIn"]]
    for i, c in enumerate(certificates, 1):
        data.append([
            str(i),
            c.get("certificate_name", ""),
            "Not Updated" if c.get("status") == "Status" else c.get("status", ""),
            c.get("deadline", ""),
            c.get("linkedin_status") or "Not Updated",
        ])
    cert_table = Table(data, colWidths=[35, 190, 85, 85, 90], repeatRows=1)
    cert_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#e9ecef")),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
    ]))
    story += [cert_table]
    doc.build(story)
    output.seek(0)
    return output.getvalue()


def make_achievement_card(student, certificates):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=70, bottomMargin=70)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("CardTitle", parent=styles["Title"], alignment=TA_CENTER, fontSize=24, leading=28)
    center = ParagraphStyle("CardCenter", parent=styles["BodyText"], alignment=TA_CENTER, fontSize=13, leading=20)

    completed = sum(1 for c in certificates if c.get("status") == "Completed")
    total = len(certificates)
    badge_name, badge_message = get_achievement_badge(completed, total)
    progress = f"{completed} / {total}" if total else "0 / 0"

    story = [
        Paragraph("DIGITAL ACHIEVEMENT CARD", title),
        Spacer(1, 25),
        Paragraph("🎓 Student Certificate Management System", center),
        Spacer(1, 20),
        Paragraph(f"<b>{student.get('name','')}</b>", center),
        Paragraph(f"Register Number: {student.get('register_no','')}", center),
        Paragraph(f"Department: {student.get('department','')}", center),
        Spacer(1, 20),
        Paragraph(f"<b>{badge_name}</b>", ParagraphStyle("Badge", parent=center, fontSize=20)),
        Spacer(1, 10),
        Paragraph(badge_message, center),
        Spacer(1, 15),
        Paragraph(f"Certificates Completed: <b>{progress}</b>", center),
        Spacer(1, 25),
        Paragraph("This card is generated by the Student Certificate Management System.", center),
    ]
    doc.build(story)
    output.seek(0)
    return output.getvalue()

# ============================================================
# CUSTOM CSS
# ============================================================

st.markdown(
    """
    <style>

    .main-title {
        font-size: 38px;
        font-weight: 700;
        text-align: center;
        margin-bottom: 5px;
    }

    .sub-title {
        text-align: center;
        font-size: 18px;
        margin-bottom: 30px;
    }

    .section-title {
        font-size: 25px;
        font-weight: 700;
        margin-top: 20px;
        margin-bottom: 15px;
    }

    .login-card {
        padding: 30px;
        border-radius: 15px;
        border: 1px solid #dddddd;
        margin-top: 20px;
        margin-bottom: 20px;
    }

    .info-card {
        padding: 20px;
        border-radius: 12px;
        border: 1px solid #dddddd;
        margin-bottom: 15px;
    }

    .footer {
        text-align: center;
        margin-top: 40px;
        padding: 20px;
        font-size: 14px;
    }

    .college-name {
        text-align: center;
        margin-top: 10px;
        margin-bottom: 20px;
    }

    .college-name h2 {
        margin-bottom: 3px;
        font-size: 28px;
        font-weight: 700;
    }

    .college-name p {
        margin-top: 0;
        font-size: 16px;
    }

    </style>
    """,
    unsafe_allow_html=True
)


# ============================================================
# HEADER
# ============================================================

st.markdown(
    """
    <div class="main-title">
        🎓 Student Certificate Management System
    </div>

    <div class="sub-title">
        Certificate Tracking & Deadline Management
    </div>
    """,
    unsafe_allow_html=True
)


# ============================================================
# SESSION STATE
# ============================================================

if "logged_in" not in st.session_state:

    st.session_state.logged_in = False


if "user_type" not in st.session_state:

    st.session_state.user_type = None


if "register_no" not in st.session_state:

    st.session_state.register_no = None


# ============================================================
# SIDEBAR
# ============================================================

with st.sidebar:

    st.image(
        "https://cdn-icons-png.flaticon.com/512/3135/3135755.png",
        width=100
    )

    st.markdown(
        "## 🎓 Certificate System"
    )

    st.markdown("---")

    if not st.session_state.logged_in:

        login_type = st.radio(
            "Login Type",
            [
                "Student Login",
                "Admin Login"
            ]
        )

    else:

        login_type = st.session_state.user_type

        if login_type == "Student":

            st.success(
                "👨‍🎓 Student Logged In"
            )

        else:

            st.success(
                "👨‍💼 Admin Logged In"
            )

        st.markdown("---")

        if st.button(
            "🚪 Logout",
            use_container_width=True
        ):

            st.session_state.logged_in = False
            st.session_state.user_type = None
            st.session_state.register_no = None

            st.rerun()

    st.markdown("---")

    st.info(
        """
        📌 Students can check their certificates,
        deadlines and upload completed certificates.

        📌 Admin can manage students,
        certificates and view analytics.
        """
    )


# ============================================================
# LOGIN SECTION
# ============================================================

if not st.session_state.logged_in:

    st.markdown(
        """
        <div class="college-name">
            <h2>RATHINAM GLOBAL DEEMED TO BE UNIVERSITY</h2>
        </div>
        """,
        unsafe_allow_html=True
    )

    # ========================================================
    # STUDENT LOGIN
    # ========================================================

    if login_type == "Student Login":

        st.markdown(
            '<div class="login-card">',
            unsafe_allow_html=True
        )

        st.markdown(
            "## 👨‍🎓 Student Login"
        )

        register_no = st.text_input(
            "Register Number",
            placeholder="Enter your register number"
        )

        password = st.text_input(
            "Password",
            type="password",
            placeholder="Enter your password"
        )

        login_button = st.button(
            "🔐 Login",
            use_container_width=True
        )

        st.markdown(
            "</div>",
            unsafe_allow_html=True
        )

        if login_button:

            if not register_no or not password:

                st.warning(
                    "⚠️ Please enter register number and password."
                )

            else:

                student = get_student(
                    register_no.strip()
                )

                if student:

                    stored_password = student["password"]

                    if verify_password(
                        password,
                        stored_password
                    ):

                        st.session_state.logged_in = True
                        st.session_state.user_type = "Student"
                        st.session_state.register_no = (
                            register_no.strip()
                        )

                        st.rerun()

                    elif stored_password == password:

                        new_hashed_password = (
                            hash_password(password)
                        )

                        (
                            supabase
                            .table("students")
                            .update(
                                {
                                    "password":
                                    new_hashed_password
                                }
                            )
                            .eq(
                                "register_no",
                                register_no.strip()
                            )
                            .execute()
                        )

                        st.session_state.logged_in = True
                        st.session_state.user_type = "Student"
                        st.session_state.register_no = (
                            register_no.strip()
                        )

                        st.rerun()

                    else:

                        st.error(
                            "❌ Invalid password."
                        )

                else:

                    st.error(
                        "❌ Student not found."
                    )


    # ========================================================
    # ADMIN LOGIN
    # ========================================================

    else:

        st.markdown(
            '<div class="login-card">',
            unsafe_allow_html=True
        )

        st.markdown(
            "## 👨‍💼 Admin Login"
        )

        admin_username = st.text_input(
            "Username"
        )

        admin_password = st.text_input(
            "Password",
            type="password"
        )

        admin_login_button = st.button(
            "🔐 Admin Login",
            use_container_width=True
        )

        st.markdown(
            "</div>",
            unsafe_allow_html=True
        )

        if admin_login_button:

            if (
                admin_username == "admin"
                and admin_password == "admin123"
            ):

                st.session_state.logged_in = True
                st.session_state.user_type = "Admin"

                st.rerun()

            else:

                st.error(
                    "❌ Invalid admin credentials."
                )


# ============================================================
# STUDENT DASHBOARD
# ============================================================

if (
    st.session_state.logged_in
    and st.session_state.user_type == "Student"
):

    register_no = st.session_state.register_no

    student = get_student(
        register_no
    )

    certificates = get_certificates(
        register_no
    )

    st.markdown(
        '<div class="section-title">👨‍🎓 Student Dashboard</div>',
        unsafe_allow_html=True
    )

    if student:

        col1, col2, col3 = st.columns(3)

        with col1:

            st.info(
                f"""
                **👤 Name**

                {student["name"]}
                """
            )

        with col2:

            st.info(
                f"""
                **🆔 Register Number**

                {student["register_no"]}
                """
            )

        with col3:

            st.info(
                f"""
                **🏫 Department**

                {student["department"]}
                """
            )

        total_certificates = len(certificates)

        completed_certificates = sum(
            1
            for c in certificates
            if c["status"] == "Completed"
        )

        pending_certificates = sum(
            1
            for c in certificates
            if c["status"] == "Pending"
        )

        not_updated_certificates = sum(
            1
            for c in certificates
            if c["status"] == "Status"
        )

        col1, col2, col3, col4 = st.columns(4)

        with col1:

            st.metric(
                "📜 Total",
                total_certificates
            )

        with col2:

            st.metric(
                "✅ Completed",
                completed_certificates
            )

        with col3:

            st.metric(
                "⏳ Pending",
                pending_certificates
            )

        with col4:

            st.metric(
                "❓ Not Updated",
                not_updated_certificates
            )

        if total_certificates > 0:

            progress = (
                completed_certificates
                / total_certificates
            )

        else:

            progress = 0

        st.markdown(
            "### 📈 Certificate Progress"
        )

        st.progress(
            progress
        )

        st.write(
            f"**{completed_certificates} / "
            f"{total_certificates} certificates completed "
            f"({progress * 100:.0f}%)**"
        )

        # ====================================================
        # ACHIEVEMENT BADGE
        # ====================================================

        badge_name, badge_message = get_achievement_badge(
            completed_certificates,
            total_certificates
        )

        st.markdown(
            "### 🏆 My Achievement"
        )

        badge_col1, badge_col2 = st.columns(
            [1, 3]
        )

        with badge_col1:

            st.markdown(
                f"""
                <div style="
                    text-align:center;
                    padding:20px;
                    border-radius:15px;
                    border:2px solid #dddddd;
                    font-size:50px;
                ">
                    {badge_name.split(" ")[0]}
                </div>
                """,
                unsafe_allow_html=True
            )

        with badge_col2:

            st.success(
                f"**{badge_name}**\n\n"
                f"{badge_message}"
            )

        if total_certificates > 0:

            chart_data = pd.DataFrame(
                {
                    "Status": [
                        "Completed",
                        "Pending",
                        "Not Updated"
                    ],
                    "Count": [
                        completed_certificates,
                        pending_certificates,
                        not_updated_certificates
                    ]
                }
            )

            chart_data = chart_data[
                chart_data["Count"] > 0
            ]

            if not chart_data.empty:

                st.markdown(
                    "### 🥧 Certificate Status"
                )

                fig = px.pie(
                    chart_data,
                    names="Status",
                    values="Count",
                    title="My Certificate Status"
                )

                st.plotly_chart(
                    fig,
                    use_container_width=True
                )

        # ====================================================
        # LINKEDIN PROFILE
        # ====================================================

        st.markdown(
            "### 🔗 LinkedIn Profile"
        )

        current_linkedin_profile = get_linkedin_profile_url(
            student
        )

        linkedin_profile_url = st.text_input(
            "LinkedIn Profile URL",
            value=current_linkedin_profile,
            placeholder="https://www.linkedin.com/in/your-profile",
            key="linkedin_profile_url_input"
        )

        if st.button(
            "💾 Save LinkedIn Profile",
            use_container_width=True
        ):
            profile_url = linkedin_profile_url.strip()

            # Automatically add https:// if the student does not enter it
            if profile_url and not profile_url.startswith(("http://", "https://")):
                profile_url = "https://" + profile_url

            # Validate LinkedIn profile URL
            if profile_url and not (
                profile_url.startswith("https://www.linkedin.com/in/")
                or profile_url.startswith("https://linkedin.com/in/")
            ):
                st.error(
                    "❌ Please enter a valid LinkedIn profile URL."
                )
            else:
                try:
                    update_linkedin_profile(
                        register_no,
                        profile_url or None
                    )
                    st.success(
                        "✅ LinkedIn profile saved successfully."
                    )
                    st.rerun()
                except Exception as e:
                    st.error(
                        f"❌ Unable to save LinkedIn profile: {e}"
                    )

        if current_linkedin_profile:
            st.markdown(
                f"🔗 [Open My LinkedIn Profile]({current_linkedin_profile})"
            )

        # ====================================================
        # LINKEDIN CERTIFICATE UPDATES
        # ====================================================

        completed_for_linkedin = [
            c for c in certificates
            if c["status"] == "Completed"
        ]

        st.markdown(
            "### 📢 LinkedIn Certificate Updates"
        )

        if current_linkedin_profile:
            st.caption(
                "Your LinkedIn profile is saved above. Mark each completed certificate that you have posted on LinkedIn."
            )

            if completed_for_linkedin:
                for certificate in completed_for_linkedin:
                    linkedin_status = get_linkedin_status(
                        certificate
                    )

                    col1, col2 = st.columns([4, 1])

                    with col1:
                        if linkedin_status == "Verified":
                            st.success(
                                f"📜 {certificate['certificate_name']} — ✅ Verified"
                            )
                        elif linkedin_status == "Submitted":
                            st.info(
                                f"📜 {certificate['certificate_name']} — 🔗 Posted / Submitted"
                            )
                        else:
                            st.warning(
                                f"📜 {certificate['certificate_name']} — ❌ Not Updated"
                            )

                    with col2:
                        if linkedin_status in [
                            "Not Updated",
                            None
                        ]:
                            if st.button(
                                "Mark Posted",
                                key=f"linkedin_posted_{certificate['id']}"
                            ):
                                try:
                                    update_linkedin_status(
                                        certificate["id"],
                                        "Submitted"
                                    )
                                    st.rerun()
                                except Exception as e:
                                    st.error(
                                        f"❌ Error: {e}"
                                    )
                        else:
                            if st.button(
                                "Undo",
                                key=f"linkedin_undo_{certificate['id']}"
                            ):
                                try:
                                    update_linkedin_status(
                                        certificate["id"],
                                        "Not Updated"
                                    )
                                    st.rerun()
                                except Exception as e:
                                    st.error(
                                        f"❌ Error: {e}"
                                    )
            else:
                st.info(
                    "📭 Complete a certificate first to track its LinkedIn update."
                )
        else:
            st.info(
                "🔗 Add your LinkedIn profile URL above to start LinkedIn tracking."
            )

        st.markdown(
            "### 📋 Certificate Details"
        )

        if certificates:

            for certificate in certificates:

                certificate_name = (
                    certificate["certificate_name"]
                )

                status = certificate["status"]

                deadline = certificate["deadline"]

                if status == "Completed":

                    st.success(
                        f"📜 {certificate_name} | "
                        f"Status: ✅ Completed | "
                        f"Deadline: {deadline}"
                    )

                elif status == "Pending":

                    st.warning(
                        f"📜 {certificate_name} | "
                        f"Status: ⏳ Pending | "
                        f"Deadline: {deadline}"
                    )

                else:

                    st.info(
                        f"📜 {certificate_name} | "
                        f"Status: ❓ Not Updated | "
                        f"Deadline: {deadline}"
                    )

                if status == "Completed":

                    answer = st.radio(
                        "Did you complete this certificate?",
                        ["Yes", "No"],
                        index=0,
                        key=f"answer_{certificate['id']}"
                    )

                elif status == "Pending":

                    answer = st.radio(
                        "Did you complete this certificate?",
                        ["Yes", "No"],
                        index=1,
                        key=f"answer_{certificate['id']}"
                    )

                else:

                    answer = st.radio(
                        "Did you complete this certificate?",
                        ["Yes", "No"],
                        index=None,
                        key=f"answer_{certificate['id']}"
                    )

                if answer:

                    new_status = (
                        "Completed"
                        if answer == "Yes"
                        else "Pending"
                    )

                    if new_status != status:

                        (
                            supabase
                            .table("certificates")
                            .update(
                                {
                                    "status":
                                    new_status
                                }
                            )
                            .eq(
                                "id",
                                certificate["id"]
                            )
                            .execute()
                        )

                        st.success(
                            "✅ Certificate status updated."
                        )

                        st.rerun()

                try:

                    deadline_date = datetime.strptime(
                        deadline,
                        "%Y-%m-%d"
                    ).date()

                    today = datetime.today().date()

                    days_left = (
                        deadline_date - today
                    ).days

                    if status != "Completed":

                        if days_left < 0:

                            st.error(
                                "🚨 Deadline expired!"
                            )

                        elif days_left <= 3:

                            st.error(
                                f"⚠️ Only {days_left} "
                                f"day(s) left!"
                            )

                        elif days_left <= 7:

                            st.warning(
                                f"⏰ {days_left} "
                                f"day(s) remaining."
                            )

                except:

                    pass

                if status == "Completed":

                    st.markdown(
                        "**📤 Upload Completed Certificate**"
                    )

                    uploaded_file = st.file_uploader(
                        "Choose certificate file",
                        type=[
                            "pdf",
                            "png",
                            "jpg",
                            "jpeg"
                        ],
                        key=f"upload_{certificate['id']}"
                    )

                    if uploaded_file:

                        if st.button(
                            "📤 Upload Certificate",
                            key=f"upload_btn_{certificate['id']}"
                        ):

                            file_path = storage_file_path(
                                register_no,
                                certificate_name,
                                uploaded_file.name
                            )

                            file_bytes = (
                                uploaded_file.getvalue()
                            )

                            content_type = (
                                uploaded_file.type
                                or "application/octet-stream"
                            )

                            try:

                                (
                                    supabase
                                    .storage
                                    .from_(STORAGE_BUCKET)
                                    .upload(
                                        file_path,
                                        file_bytes,
                                        {
                                            "content-type":
                                            content_type,
                                            "upsert": "true"
                                        }
                                    )
                                )

                                st.success(
                                    "✅ Certificate uploaded successfully."
                                )

                            except Exception as e:

                                st.error(
                                    f"❌ Upload failed: {e}"
                                )

        else:

            st.info(
                "📭 No certificates assigned yet."
            )


# ============================================================
# ADMIN DASHBOARD
# ============================================================

if (
    st.session_state.logged_in
    and st.session_state.user_type == "Admin"
):

    st.markdown(
        '<div class="section-title">👨‍💼 Admin Dashboard</div>',
        unsafe_allow_html=True
    )

    all_students = get_students()

    all_certificates = get_certificates()

    all_students = sorted(
        all_students,
        key=lambda x:
        get_numeric_register_number(
            x["register_no"]
        )
    )

    all_certificates = sorted(
        all_certificates,
        key=lambda x:
        get_numeric_register_number(
            x["register_no"]
        )
    )

    total_students = len(
        all_students
    )

    total_completed = 0
    total_pending = 0
    total_not_updated = 0

    for student in all_students:
        student_register = student["register_no"]
        student_certificates = [
            c for c in all_certificates
            if c["register_no"] == student_register
        ]

        if not student_certificates:
            total_not_updated += 1
            continue

        statuses = [c["status"] for c in student_certificates]
        if all(status == "Completed" for status in statuses):
            total_completed += 1
        elif any(status == "Pending" for status in statuses):
            total_pending += 1
        else:
            total_not_updated += 1

    col1, col2, col3, col4 = st.columns(4)

    with col1:

        st.metric(
            "👨‍🎓 Total Students",
            total_students
        )

    with col2:

        st.metric(
            "✅ Completed",
            total_completed
        )

    with col3:

        st.metric(
            "⏳ Pending",
            total_pending
        )

    with col4:

        st.metric(
            "❓ Not Updated",
            total_not_updated
        )

    st.markdown("---")

    # ========================================================
    # ADMIN MENU
    # ========================================================

    admin_menu = st.selectbox(
        "Admin Menu",
        [
            "Add New Student",
            "Remove Student",
            "Add Certificate",
            "Remove Certificate",
            "Student List",
            "Student Search & Filter",
            "Certificate Overview",
            "Certificate Analytics",
            "Update Certificate",
            "LinkedIn Updates",
            "Uploaded Certificates"
        ]
    )


    # ========================================================
    # ADD NEW STUDENT
    # ========================================================

    if admin_menu == "Add New Student":

        st.subheader(
            "➕ Add New Student"
        )

        if all_students:

            st.info(
                f"👨‍🎓 {len(all_students)} student(s) already registered."
            )

        new_register_no = st.text_input(
            "Register Number"
        )

        new_name = st.text_input(
            "Student Name"
        )

        new_department = st.text_input(
            "Department"
        )

        new_password = st.text_input(
            "Student Password",
            type="password"
        )

        if st.button(
            "➕ Add Student"
        ):

            if (
                new_register_no
                and new_name
                and new_department
                and new_password
            ):

                existing_student = get_student(
                    new_register_no.strip()
                )

                if existing_student:

                    st.error(
                        "❌ Student already exists."
                    )

                else:

                    hashed_password = hash_password(
                        new_password
                    )

                    try:

                        (
                            supabase
                            .table("students")
                            .insert(
                                {
                                    "register_no":
                                    new_register_no.strip(),

                                    "name":
                                    new_name.strip(),

                                    "department":
                                    new_department.strip(),

                                    "password":
                                    hashed_password
                                }
                            )
                            .execute()
                        )

                        st.success(
                            "✅ Student added successfully."
                        )

                        st.rerun()

                    except Exception as e:

                        st.error(
                            f"❌ Error: {e}"
                        )

            else:

                st.warning(
                    "⚠️ Please fill all fields."
                )


    # ========================================================
    # REMOVE STUDENT
    # ========================================================

    elif admin_menu == "Remove Student":

        st.subheader(
            "🗑️ Remove Student"
        )

        if all_students:

            student_options = [
                f'{s["register_no"]} - {s["name"]}'
                for s in all_students
            ]

            selected_student = st.selectbox(
                "Select Student",
                student_options
            )

            selected_register = (
                selected_student
                .split(" - ")[0]
            )

            if st.button(
                "🗑️ Remove Student"
            ):

                try:

                    (
                        supabase
                        .table("certificates")
                        .delete()
                        .eq(
                            "register_no",
                            selected_register
                        )
                        .execute()
                    )

                    (
                        supabase
                        .table("students")
                        .delete()
                        .eq(
                            "register_no",
                            selected_register
                        )
                        .execute()
                    )

                    st.success(
                        "✅ Student removed successfully."
                    )

                    st.rerun()

                except Exception as e:

                    st.error(
                        f"❌ Error: {e}"
                    )

        else:

            st.info(
                "📭 No students available."
            )


    # ========================================================
    # ADD CERTIFICATE
    # ========================================================

    elif admin_menu == "Add Certificate":

        st.subheader(
            "📜 Add Certificate"
        )

        if all_students:

            # ------------------------------------------------
            # ASSIGN TYPE
            # ------------------------------------------------

            assign_type = st.selectbox(
                "Assign Certificate To",
                [
                    "All Students",
                    "Individual Student"
                ]
            )

            # ------------------------------------------------
            # INDIVIDUAL STUDENT
            # ------------------------------------------------

            selected_register = None

            if assign_type == "Individual Student":

                student_options = [
                    f'{s["register_no"]} - {s["name"]}'
                    for s in all_students
                ]

                selected_student = st.selectbox(
                    "Select Student",
                    student_options
                )

                selected_register = (
                    selected_student
                    .split(" - ")[0]
                )

            else:

                st.info(
                    f"📢 This certificate will be assigned "
                    f"to all {len(all_students)} students."
                )

            # ------------------------------------------------
            # CERTIFICATE DETAILS
            # ------------------------------------------------

            certificate_name = st.text_input(
                "Certificate Name"
            )

            certificate_status = st.selectbox(
                "Initial Status",
                [
                    "Status",
                    "Pending",
                    "Completed"
                ]
            )

            certificate_deadline = st.date_input(
                "Deadline"
            )

            if st.button(
                "➕ Add Certificate"
            ):

                if certificate_name.strip():

                    certificate_name_clean = (
                        certificate_name
                        .strip()
                    )

                    deadline_value = (
                        certificate_deadline
                        .strftime("%Y-%m-%d")
                    )

                    try:

                        # ====================================
                        # ALL STUDENTS
                        # ====================================

                        if assign_type == "All Students":

                            # Get latest certificate data
                            latest_certificates = (
                                get_certificates()
                            )

                            # Existing certificate pairs
                            existing_pairs = set()

                            for certificate in latest_certificates:

                                existing_register = str(
                                    certificate[
                                        "register_no"
                                    ]
                                ).strip()

                                existing_name = str(
                                    certificate[
                                        "certificate_name"
                                    ]
                                ).strip().lower()

                                existing_pairs.add(
                                    (
                                        existing_register,
                                        existing_name
                                    )
                                )

                            rows_to_insert = []

                            for student in all_students:

                                student_register = (
                                    str(
                                        student[
                                            "register_no"
                                        ]
                                    ).strip()
                                )

                                certificate_key = (
                                    student_register,
                                    certificate_name_clean.lower()
                                )

                                # Avoid duplicate
                                if certificate_key not in existing_pairs:

                                    rows_to_insert.append(
                                        {
                                            "register_no":
                                            student_register,

                                            "certificate_name":
                                            certificate_name_clean,

                                            "status":
                                            certificate_status,

                                            "deadline":
                                            deadline_value
                                        }
                                    )

                            if rows_to_insert:

                                (
                                    supabase
                                    .table("certificates")
                                    .insert(
                                        rows_to_insert
                                    )
                                    .execute()
                                )

                                st.success(
                                    f"✅ Certificate added successfully "
                                    f"to {len(rows_to_insert)} student(s)."
                                )

                            else:

                                st.warning(
                                    "⚠️ This certificate already exists "
                                    "for all students."
                                )

                            st.rerun()

                        # ====================================
                        # INDIVIDUAL STUDENT
                        # ====================================

                        else:

                            latest_certificates = (
                                get_certificates(
                                    selected_register
                                )
                            )

                            duplicate_exists = any(
                                str(
                                    certificate[
                                        "certificate_name"
                                    ]
                                ).strip().lower()
                                ==
                                certificate_name_clean.lower()
                                for certificate
                                in latest_certificates
                            )

                            if duplicate_exists:

                                st.warning(
                                    "⚠️ This certificate already "
                                    "exists for this student."
                                )

                            else:

                                (
                                    supabase
                                    .table("certificates")
                                    .insert(
                                        {
                                            "register_no":
                                            selected_register,

                                            "certificate_name":
                                            certificate_name_clean,

                                            "status":
                                            certificate_status,

                                            "deadline":
                                            deadline_value
                                        }
                                    )
                                    .execute()
                                )

                                st.success(
                                    "✅ Certificate added successfully."
                                )

                                st.rerun()

                    except Exception as e:

                        st.error(
                            f"❌ Error: {e}"
                        )

                else:

                    st.warning(
                        "⚠️ Enter certificate name."
                    )

        else:

            st.info(
                "📭 Add students first."
            )


    # ========================================================
    # REMOVE CERTIFICATE
    # ========================================================

    elif admin_menu == "Remove Certificate":

        st.subheader(
            "🗑️ Remove Certificate"
        )

        if all_certificates:

            certificate_options = [
                f'{c["id"]} - '
                f'{c["register_no"]} - '
                f'{c["certificate_name"]}'
                for c in all_certificates
            ]

            selected_certificate = st.selectbox(
                "Select Certificate",
                certificate_options
            )

            selected_id = int(
                selected_certificate
                .split(" - ")[0]
            )

            if st.button(
                "🗑️ Remove Certificate"
            ):

                try:

                    (
                        supabase
                        .table("certificates")
                        .delete()
                        .eq(
                            "id",
                            selected_id
                        )
                        .execute()
                    )

                    st.success(
                        "✅ Certificate removed successfully."
                    )

                    st.rerun()

                except Exception as e:

                    st.error(
                        f"❌ Error: {e}"
                    )

        else:

            st.info(
                "📭 No certificates available."
            )


    # ========================================================
    # STUDENT LIST
    # ========================================================

    elif admin_menu == "Student List":

        st.subheader(
            "👨‍🎓 Student List"
        )

        if all_students:

            student_rows = []

            for student in all_students:

                student_rows.append(
                    {
                        "Register Number":
                        student["register_no"],

                        "Name":
                        student["name"],

                        "Department":
                        student["department"]
                    }
                )

            student_df = pd.DataFrame(
                student_rows
            )

            st.dataframe(
                student_df,
                use_container_width=True,
                hide_index=True
            )

        else:

            st.info(
                "📭 No students available."
            )


    # ========================================================
    # STUDENT SEARCH & FILTER
    # ========================================================

    elif admin_menu == "Student Search & Filter":

        st.subheader(
            "🔍 Student Search & Filter"
        )

        st.write(
            "Search students and filter their certificate status."
        )

        if all_students:

            search_text = st.text_input(
                "🔍 Search by Register Number or Student Name",
                placeholder="Example: RCAS2026BAI173 or Pooja"
            )

            department_values = sorted(
                list(
                    set(
                        str(student["department"])
                        for student in all_students
                        if student.get("department")
                    )
                )
            )

            department_options = [
                "All Departments"
            ] + department_values

            certificate_values = sorted(
                list(
                    set(
                        str(certificate["certificate_name"])
                        for certificate in all_certificates
                        if certificate.get("certificate_name")
                    )
                )
            )

            certificate_options = [
                "All Certificates"
            ] + certificate_values

            status_options = [
                "All Statuses",
                "Completed",
                "Pending",
                "Not Updated"
            ]

            filter_col1, filter_col2, filter_col3 = (
                st.columns(3)
            )

            with filter_col1:

                selected_department = st.selectbox(
                    "🏫 Department",
                    department_options
                )

            with filter_col2:

                selected_certificate = st.selectbox(
                    "📜 Certificate",
                    certificate_options
                )

            with filter_col3:

                selected_status = st.selectbox(
                    "📊 Certificate Status",
                    status_options
                )

            filtered_students = []

            for student in all_students:

                register = str(
                    student["register_no"]
                )

                name = str(
                    student["name"]
                )

                department = str(
                    student["department"]
                )

                if search_text:

                    search_lower = (
                        search_text
                        .strip()
                        .lower()
                    )

                    if (
                        search_lower
                        not in register.lower()
                        and search_lower
                        not in name.lower()
                    ):

                        continue

                if (
                    selected_department
                    != "All Departments"
                    and department
                    != selected_department
                ):

                    continue

                student_certificates = [
                    certificate
                    for certificate in all_certificates
                    if certificate[
                        "register_no"
                    ] == register
                ]

                if (
                    selected_certificate
                    != "All Certificates"
                ):

                    matching_certificates = [
                        certificate
                        for certificate
                        in student_certificates
                        if certificate[
                            "certificate_name"
                        ]
                        == selected_certificate
                    ]

                    if not matching_certificates:

                        continue

                    student_certificates = (
                        matching_certificates
                    )

                if (
                    selected_status
                    != "All Statuses"
                ):

                    status_match = False

                    for certificate in student_certificates:

                        certificate_status = (
                            certificate["status"]
                        )

                        if (
                            selected_status
                            == "Completed"
                            and certificate_status
                            == "Completed"
                        ):

                            status_match = True

                        elif (
                            selected_status
                            == "Pending"
                            and certificate_status
                            == "Pending"
                        ):

                            status_match = True

                        elif (
                            selected_status
                            == "Not Updated"
                            and certificate_status
                            == "Status"
                        ):

                            status_match = True

                    if not status_match:

                        continue

                total = len(
                    student_certificates
                )

                completed = sum(
                    1
                    for certificate
                    in student_certificates
                    if certificate[
                        "status"
                    ] == "Completed"
                )

                pending = sum(
                    1
                    for certificate
                    in student_certificates
                    if certificate[
                        "status"
                    ] == "Pending"
                )

                not_updated = sum(
                    1
                    for certificate
                    in student_certificates
                    if certificate[
                        "status"
                    ] == "Status"
                )

                filtered_students.append(
                    {
                        "Register Number":
                        register,

                        "Student Name":
                        name,

                        "Department":
                        department,

                        "Total Certificates":
                        total,

                        "Completed":
                        completed,

                        "Pending":
                        pending,

                        "Not Updated":
                        not_updated
                    }
                )

            st.markdown("---")

            st.markdown(
                f"### 📋 Search Results: "
                f"{len(filtered_students)} Student(s)"
            )

            if filtered_students:

                filtered_df = pd.DataFrame(
                    filtered_students
                )

                st.dataframe(
                    filtered_df,
                    use_container_width=True,
                    hide_index=True
                )

                result_col1, result_col2, result_col3, result_col4 = (
                    st.columns(4)
                )

                result_total = sum(
                    row["Total Certificates"]
                    for row in filtered_students
                )

                result_completed = sum(
                    row["Completed"]
                    for row in filtered_students
                )

                result_pending = sum(
                    row["Pending"]
                    for row in filtered_students
                )

                result_not_updated = sum(
                    row["Not Updated"]
                    for row in filtered_students
                )

                with result_col1:

                    st.metric(
                        "📜 Total Certificates",
                        result_total
                    )

                with result_col2:

                    st.metric(
                        "✅ Completed",
                        result_completed
                    )

                with result_col3:

                    st.metric(
                        "⏳ Pending",
                        result_pending
                    )

                with result_col4:

                    st.metric(
                        "❓ Not Updated",
                        result_not_updated
                    )

                st.markdown(
                    "### 👨‍🎓 Student Certificate Details"
                )

                for row in filtered_students:

                    student_register = (
                        row["Register Number"]
                    )

                    student_name = (
                        row["Student Name"]
                    )

                    student_department = (
                        row["Department"]
                    )

                    student_certificate_data = [
                        certificate
                        for certificate
                        in all_certificates
                        if certificate[
                            "register_no"
                        ] == student_register
                    ]

                    with st.expander(
                        f"📁 {student_name} | "
                        f"🆔 {student_register}"
                    ):

                        detail_col1, detail_col2, detail_col3 = (
                            st.columns(3)
                        )

                        with detail_col1:

                            st.write(
                                f"**👤 Name:** "
                                f"{student_name}"
                            )

                        with detail_col2:

                            st.write(
                                f"**🆔 Register Number:** "
                                f"{student_register}"
                            )

                        with detail_col3:

                            st.write(
                                f"**🏫 Department:** "
                                f"{student_department}"
                            )

                        st.markdown("---")

                        if student_certificate_data:

                            for certificate in student_certificate_data:

                                certificate_name = (
                                    certificate[
                                        "certificate_name"
                                    ]
                                )

                                certificate_status = (
                                    certificate[
                                        "status"
                                    ]
                                )

                                deadline = (
                                    certificate[
                                        "deadline"
                                    ]
                                )

                                if certificate_status == "Completed":

                                    st.success(
                                        f"📜 **{certificate_name}** | "
                                        f"✅ Completed | "
                                        f"Deadline: {deadline}"
                                    )

                                elif certificate_status == "Pending":

                                    st.warning(
                                        f"📜 **{certificate_name}** | "
                                        f"⏳ Pending | "
                                        f"Deadline: {deadline}"
                                    )

                                else:

                                    st.info(
                                        f"📜 **{certificate_name}** | "
                                        f"❓ Not Updated | "
                                        f"Deadline: {deadline}"
                                    )

                        else:

                            st.info(
                                "📭 No certificates assigned."
                            )

            else:

                st.warning(
                    "🔎 No students match the selected filters."
                )

        else:

            st.info(
                "📭 No students available."
            )


    # ========================================================
    # CERTIFICATE OVERVIEW
    # ========================================================

    elif admin_menu == "Certificate Overview":

        st.subheader("📂 Certificate Overview")
        st.write("View certificate status and preview/download uploaded certificate files.")

        if all_students:
            for student in all_students:
                student_register = student["register_no"]
                student_name = student["name"]
                student_certificates = [
                    c for c in all_certificates
                    if c["register_no"] == student_register
                ]

                with st.expander(
                    f"📁 {student_name} | 🆔 {student_register} | 📜 {len(student_certificates)} Certificate(s)"
                ):
                    st.write(f"**👤 Name:** {student_name}")
                    st.write(f"**🆔 Register Number:** {student_register}")
                    st.write(f"**🏫 Department:** {student['department']}")
                    st.markdown("---")

                    if student_certificates:
                        for certificate in student_certificates:
                            certificate_name = certificate["certificate_name"]
                            status = certificate["status"]
                            deadline = certificate["deadline"]

                            if status == "Completed":
                                st.success(f"📜 **{certificate_name}** | ✅ Completed | Deadline: {deadline}")
                            elif status == "Pending":
                                st.warning(f"📜 **{certificate_name}** | ⏳ Pending | Deadline: {deadline}")
                            else:
                                st.info(f"📜 **{certificate_name}** | ❓ Not Updated | Deadline: {deadline}")

                            uploaded_files = get_uploaded_certificate_files(student_register, certificate_name)
                            if uploaded_files:
                                st.write("**📎 Uploaded Certificate:**")
                                for file_name in uploaded_files:
                                    try:
                                        file_bytes = download_storage_file(file_name)
                                        ext = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""
                                        view_col, download_col = st.columns(2)
                                        with view_col:
                                            if st.button("👁️ Preview", key=f"overview_preview_{certificate['id']}_{file_name}"):
                                                if ext == "pdf":
                                                    st.pdf(file_bytes)
                                                elif ext in ["png", "jpg", "jpeg"]:
                                                    st.image(file_bytes, caption=file_name, use_container_width=True)
                                                else:
                                                    st.info("Preview is not available for this file type.")
                                        with download_col:
                                            st.download_button(
                                                "⬇️ Download",
                                                data=file_bytes,
                                                file_name=file_name,
                                                key=f"overview_download_{certificate['id']}_{file_name}"
                                            )
                                    except Exception as e:
                                        st.error(f"❌ Unable to open {file_name}: {e}")
                            else:
                                st.caption("📭 No uploaded file for this certificate yet.")
                    else:
                        st.info("📭 No certificates assigned.")
        else:
            st.info("📭 No students available.")


    # ========================================================
    # REPORTS
    # ========================================================

    elif admin_menu in ["Excel Reports", "PDF Reports", "Digital Achievement Card"]:

        if admin_menu == "Excel Reports":
            st.subheader("📊 Excel Student Reports")
            report_title = "Download Excel report for a student."
        elif admin_menu == "PDF Reports":
            st.subheader("📄 PDF Student Reports")
            report_title = "Download PDF report for a student."
        else:
            st.subheader("🏆 Digital Achievement Card")
            report_title = "Generate a digital achievement card for a student."

        st.write(report_title)

        if all_students:
            student_options = [
                f'{s["register_no"]} - {s["name"]}'
                for s in all_students
            ]
            selected_report_student = st.selectbox(
                "Select Student",
                student_options,
                key=f"report_student_{admin_menu}"
            )
            selected_report_register = selected_report_student.split(" - ")[0]
            report_student = get_student(selected_report_register)
            report_certificates = get_certificates(selected_report_register)

            if report_student:
                completed = sum(1 for c in report_certificates if c.get("status") == "Completed")
                pending = sum(1 for c in report_certificates if c.get("status") == "Pending")
                not_updated = sum(1 for c in report_certificates if c.get("status") == "Status")

                c1, c2, c3, c4 = st.columns(4)
                c1.metric("📜 Total", len(report_certificates))
                c2.metric("✅ Completed", completed)
                c3.metric("⏳ Pending", pending)
                c4.metric("❓ Not Updated", not_updated)

                if admin_menu == "Excel Reports":
                    excel_bytes = make_excel_report(report_student, report_certificates)
                    st.download_button(
                        "⬇️ Download Excel Report",
                        data=excel_bytes,
                        file_name=f"{selected_report_register}_Certificate_Report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                elif admin_menu == "PDF Reports":
                    pdf_bytes = make_pdf_report(report_student, report_certificates)
                    st.download_button(
                        "⬇️ Download PDF Report",
                        data=pdf_bytes,
                        file_name=f"{selected_report_register}_Certificate_Report.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                    with st.expander("👁️ Preview Report"):
                        st.pdf(pdf_bytes)
                else:
                    card_bytes = make_achievement_card(report_student, report_certificates)
                    st.download_button(
                        "⬇️ Download Digital Achievement Card",
                        data=card_bytes,
                        file_name=f"{selected_report_register}_Achievement_Card.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                    with st.expander("👁️ Preview Achievement Card"):
                        st.pdf(card_bytes)
        else:
            st.info("📭 No students available.")


    # ========================================================
    # CERTIFICATE ANALYTICS
    # ========================================================

    elif admin_menu == "Certificate Analytics":

        st.subheader(
            "📊 Certificate Analytics"
        )

        st.write(
            "View certificate-wise completion status "
            "with student names and register numbers."
        )

        if all_certificates:

            student_lookup = {}

            for student in all_students:

                student_lookup[
                    student["register_no"]
                ] = student

            analytics_rows = []

            for certificate in all_certificates:

                certificate_name = (
                    certificate[
                        "certificate_name"
                    ]
                    .strip()
                )

                register_no = (
                    certificate[
                        "register_no"
                    ]
                )

                student = student_lookup.get(
                    register_no
                )

                if student:

                    student_name = (
                        student["name"]
                    )

                else:

                    student_name = "Unknown Student"

                status = certificate[
                    "status"
                ]

                if status == "Completed":

                    status_name = "Completed"

                elif status == "Pending":

                    status_name = "Pending"

                else:

                    status_name = "Not Updated"

                analytics_rows.append(
                    {
                        "Certificate":
                        certificate_name,

                        "Register Number":
                        register_no,

                        "Student Name":
                        student_name,

                        "Status":
                        status_name
                    }
                )

            analytics_df = pd.DataFrame(
                analytics_rows
            )

            summary_df = (
                analytics_df
                .groupby(
                    [
                        "Certificate",
                        "Status"
                    ]
                )
                .size()
                .unstack(
                    fill_value=0
                )
                .reset_index()
            )

            for column in [
                "Completed",
                "Pending",
                "Not Updated"
            ]:

                if column not in summary_df.columns:

                    summary_df[column] = 0

            summary_df["Total"] = (
                summary_df["Completed"]
                + summary_df["Pending"]
                + summary_df["Not Updated"]
            )

            summary_df = summary_df[
                [
                    "Certificate",
                    "Completed",
                    "Pending",
                    "Not Updated",
                    "Total"
                ]
            ]

            st.markdown(
                "### 📋 Certificate-wise Status"
            )

            st.dataframe(
                summary_df,
                use_container_width=True,
                hide_index=True
            )

            analytics_completed = int(
                summary_df[
                    "Completed"
                ].sum()
            )

            analytics_pending = int(
                summary_df[
                    "Pending"
                ].sum()
            )

            analytics_not_updated = int(
                summary_df[
                    "Not Updated"
                ].sum()
            )

            analytics_total = int(
                summary_df[
                    "Total"
                ].sum()
            )

            st.markdown(
                "### 📈 Overall Certificate Statistics"
            )

            col1, col2, col3, col4 = st.columns(4)

            with col1:

                st.metric(
                    "📜 Total",
                    analytics_total
                )

            with col2:

                st.metric(
                    "✅ Completed",
                    analytics_completed
                )

            with col3:

                st.metric(
                    "⏳ Pending",
                    analytics_pending
                )

            with col4:

                st.metric(
                    "❓ Not Updated",
                    analytics_not_updated
                )

            st.markdown(
                "### 📊 Certificate-wise Comparison"
            )

            chart_df = summary_df.melt(
                id_vars=[
                    "Certificate"
                ],
                value_vars=[
                    "Completed",
                    "Pending",
                    "Not Updated"
                ],
                var_name="Status",
                value_name="Students"
            )

            fig = px.bar(
                chart_df,
                x="Certificate",
                y="Students",
                color="Status",
                barmode="group",
                title="Certificate Completion Status"
            )

            fig.update_layout(
                height=500,
                xaxis_title="Certificate",
                yaxis_title="Number of Students",
                margin=dict(
                    l=20,
                    r=20,
                    t=60,
                    b=80
                )
            )

            st.plotly_chart(
                fig,
                use_container_width=True
            )

            st.markdown(
                "### 👨‍🎓 Student Details by Certificate"
            )

            certificate_names = sorted(
                analytics_df[
                    "Certificate"
                ].unique()
            )

            for certificate_name in certificate_names:

                certificate_data = analytics_df[
                    analytics_df[
                        "Certificate"
                    ]
                    == certificate_name
                ]

                completed_students = (
                    certificate_data[
                        certificate_data[
                            "Status"
                        ] == "Completed"
                    ]
                )

                pending_students = (
                    certificate_data[
                        certificate_data[
                            "Status"
                        ] == "Pending"
                    ]
                )

                not_updated_students = (
                    certificate_data[
                        certificate_data[
                            "Status"
                        ] == "Not Updated"
                    ]
                )

                with st.expander(
                    f"📁 {certificate_name}"
                ):

                    col1, col2, col3, col4 = st.columns(4)

                    with col1:

                        st.metric(
                            "📜 Total",
                            len(certificate_data)
                        )

                    with col2:

                        st.metric(
                            "✅ Completed",
                            len(completed_students)
                        )

                    with col3:

                        st.metric(
                            "⏳ Pending",
                            len(pending_students)
                        )

                    with col4:

                        st.metric(
                            "❓ Not Updated",
                            len(not_updated_students)
                        )

                    st.markdown("---")

                    st.markdown(
                        "#### ✅ Completed Students"
                    )

                    if not completed_students.empty:

                        completed_display = (
                            completed_students[
                                [
                                    "Register Number",
                                    "Student Name"
                                ]
                            ]
                            .reset_index(
                                drop=True
                            )
                        )

                        completed_display.index = (
                            completed_display.index + 1
                        )

                        completed_display.index.name = (
                            "S.No"
                        )

                        st.dataframe(
                            completed_display,
                            use_container_width=True
                        )

                    else:

                        st.info(
                            "No students have completed this certificate."
                        )

                    st.markdown(
                        "#### ⏳ Pending Students"
                    )

                    if not pending_students.empty:

                        pending_display = (
                            pending_students[
                                [
                                    "Register Number",
                                    "Student Name"
                                ]
                            ]
                            .reset_index(
                                drop=True
                            )
                        )

                        pending_display.index = (
                            pending_display.index + 1
                        )

                        pending_display.index.name = (
                            "S.No"
                        )

                        st.dataframe(
                            pending_display,
                            use_container_width=True
                        )

                    else:

                        st.info(
                            "No pending students."
                        )

                    st.markdown(
                        "#### ❓ Not Updated Students"
                    )

                    if not not_updated_students.empty:

                        not_updated_display = (
                            not_updated_students[
                                [
                                    "Register Number",
                                    "Student Name"
                                ]
                            ]
                            .reset_index(
                                drop=True
                            )
                        )

                        not_updated_display.index = (
                            not_updated_display.index + 1
                        )

                        not_updated_display.index.name = (
                            "S.No"
                        )

                        st.dataframe(
                            not_updated_display,
                            use_container_width=True
                        )

                    else:

                        st.info(
                            "No students with Not Updated status."
                        )

        else:

            st.info(
                "📭 No certificates available for analytics."
            )


    # ========================================================
    # UPDATE CERTIFICATE
    # ========================================================

    elif admin_menu == "Update Certificate":

        st.subheader(
            "✏️ Update Certificate"
        )

        if all_certificates:

            certificate_options = [
                f'{c["id"]} - '
                f'{c["register_no"]} - '
                f'{c["certificate_name"]}'
                for c in all_certificates
            ]

            selected_certificate = st.selectbox(
                "Select Certificate",
                certificate_options
            )

            selected_id = int(
                selected_certificate
                .split(" - ")[0]
            )

            selected_certificate_data = next(
                (
                    c
                    for c in all_certificates
                    if c["id"] == selected_id
                ),
                None
            )

            if selected_certificate_data:

                current_status = (
                    selected_certificate_data[
                        "status"
                    ]
                )

                current_deadline = datetime.strptime(
                    selected_certificate_data[
                        "deadline"
                    ],
                    "%Y-%m-%d"
                ).date()

                new_status = st.selectbox(
                    "Status",
                    [
                        "Status",
                        "Pending",
                        "Completed"
                    ],
                    index=[
                        "Status",
                        "Pending",
                        "Completed"
                    ].index(
                        current_status
                    )
                    if current_status
                    in [
                        "Status",
                        "Pending",
                        "Completed"
                    ]
                    else 0
                )

                new_deadline = st.date_input(
                    "Deadline",
                    value=current_deadline
                )

                if st.button(
                    "💾 Update Certificate"
                ):

                    try:

                        (
                            supabase
                            .table("certificates")
                            .update(
                                {
                                    "status":
                                    new_status,

                                    "deadline":
                                    new_deadline.strftime(
                                        "%Y-%m-%d"
                                    )
                                }
                            )
                            .eq(
                                "id",
                                selected_id
                            )
                            .execute()
                        )

                        st.success(
                            "✅ Certificate updated successfully."
                        )

                        st.rerun()

                    except Exception as e:

                        st.error(
                            f"❌ Error: {e}"
                        )

        else:

            st.info(
                "📭 No certificates available."
            )


    # ========================================================
    # LINKEDIN UPDATES
    # ========================================================

    elif admin_menu == "LinkedIn Updates":

        st.subheader(
            "🔗 LinkedIn Updates"
        )

        student_lookup = {
            s["register_no"]: s
            for s in all_students
        }

        completed_certificates = [
            c for c in all_certificates
            if c["status"] == "Completed"
        ]

        total_completed_linkedin = len(
            completed_certificates
        )

        verified_count = sum(
            1
            for c in completed_certificates
            if get_linkedin_status(c) == "Verified"
        )

        submitted_count = sum(
            1
            for c in completed_certificates
            if get_linkedin_status(c) == "Submitted"
        )

        not_updated_count = total_completed_linkedin - (
            verified_count + submitted_count
        )

        col1, col2, col3, col4 = st.columns(4)

        with col1:
            st.metric(
                "📜 Completed Certificates",
                total_completed_linkedin
            )

        with col2:
            st.metric(
                "✅ Verified",
                verified_count
            )

        with col3:
            st.metric(
                "🔗 Posted / Submitted",
                submitted_count
            )

        with col4:
            st.metric(
                "❌ Not Updated",
                not_updated_count
            )

        if total_completed_linkedin > 0:
            linkedin_progress = (
                (verified_count + submitted_count)
                / total_completed_linkedin
            )
        else:
            linkedin_progress = 0

        st.progress(linkedin_progress)

        st.write(
            f"**{verified_count + submitted_count} / "
            f"{total_completed_linkedin} completed certificates "
            f"are marked as LinkedIn updated "
            f"({linkedin_progress * 100:.0f}%)**"
        )

        linkedin_rows = []

        for student_item in all_students:
            student_register = student_item["register_no"]

            student_completed = [
                c for c in completed_certificates
                if c["register_no"] == student_register
            ]

            if not student_completed:
                continue

            updated_for_student = sum(
                1
                for c in student_completed
                if get_linkedin_status(c)
                in ["Submitted", "Verified"]
            )

            linkedin_rows.append({
                "Register Number": student_register,
                "Student Name": student_item["name"],
                "LinkedIn Profile": get_linkedin_profile_url(
                    student_item
                ) or "Not Added",
                "Completed Certificates": len(student_completed),
                "LinkedIn Updated": updated_for_student,
                "Pending": (
                    len(student_completed)
                    - updated_for_student
                )
            })

        if linkedin_rows:
            st.markdown(
                "### 👨‍🎓 Student-wise LinkedIn Tracking"
            )

            linkedin_df = pd.DataFrame(
                linkedin_rows
            )

            st.dataframe(
                linkedin_df,
                use_container_width=True,
                hide_index=True
            )

            st.markdown(
                "### 📋 Student Details"
            )

            for student_item in all_students:
                student_register = student_item["register_no"]

                student_completed = [
                    c for c in completed_certificates
                    if c["register_no"] == student_register
                ]

                if not student_completed:
                    continue

                student_updated = sum(
                    1
                    for c in student_completed
                    if get_linkedin_status(c)
                    in ["Submitted", "Verified"]
                )

                with st.expander(
                    f"👨‍🎓 {student_item['name']} | "
                    f"🆔 {student_register} | "
                    f"🔗 {student_updated}/{len(student_completed)} Updated"
                ):

                    profile_url = get_linkedin_profile_url(
                        student_item
                    )

                    if profile_url:
                        st.markdown(
                            f"🔗 [Open LinkedIn Profile]({profile_url})"
                        )
                    else:
                        st.warning(
                            "⚠️ LinkedIn profile not added by this student."
                        )

                    for certificate in student_completed:
                        linkedin_status = get_linkedin_status(
                            certificate
                        )

                        c1, c2 = st.columns([4, 2])

                        with c1:
                            st.write(
                                f"📜 **{certificate['certificate_name']}**"
                            )

                            if linkedin_status == "Verified":
                                st.success(
                                    "Status: ✅ Verified"
                                )
                            elif linkedin_status == "Submitted":
                                st.info(
                                    "Status: 🔗 Posted / Submitted"
                                )
                            else:
                                st.warning(
                                    "Status: ❌ Not Updated"
                                )

                        with c2:
                            if linkedin_status == "Submitted":
                                if st.button(
                                    "✅ Verify",
                                    key=f"verify_linkedin_{certificate['id']}"
                                ):
                                    try:
                                        update_linkedin_status(
                                            certificate["id"],
                                            "Verified"
                                        )
                                        st.success(
                                            "LinkedIn update verified."
                                        )
                                        st.rerun()
                                    except Exception as e:
                                        st.error(
                                            f"❌ Error: {e}"
                                        )

                            if linkedin_status in [
                                "Submitted",
                                "Verified"
                            ]:
                                if st.button(
                                    "❌ Mark Not Updated",
                                    key=f"not_updated_linkedin_{certificate['id']}"
                                ):
                                    try:
                                        update_linkedin_status(
                                            certificate["id"],
                                            "Not Updated"
                                        )
                                        st.success(
                                            "LinkedIn status updated."
                                        )
                                        st.rerun()
                                    except Exception as e:
                                        st.error(
                                            f"❌ Error: {e}"
                                        )

                        st.markdown("---")
        else:
            st.info(
                "📭 No completed certificates available for LinkedIn tracking."
            )


    # ========================================================
    # UPLOADED CERTIFICATES
    # ========================================================

    elif admin_menu == "Uploaded Certificates":

        st.subheader(
            "📤 Uploaded Certificates"
        )

        try:

            storage_files = (
                supabase
                .storage
                .from_(STORAGE_BUCKET)
                .list()
            )

            if storage_files:

                for file_info in storage_files:

                    file_name = file_info.get(
                        "name"
                    )

                    if not file_name:

                        continue

                    st.markdown(
                        f"### 📄 {file_name}"
                    )

                    col1, col2 = st.columns(2)

                    with col1:

                        if st.button(
                            "👁️ View / Download",
                            key=f"view_{file_name}"
                        ):

                            try:

                                file_bytes = (
                                    supabase
                                    .storage
                                    .from_(
                                        STORAGE_BUCKET
                                    )
                                    .download(
                                        file_name
                                    )
                                )

                                st.download_button(
                                    "⬇️ Download File",
                                    data=file_bytes,
                                    file_name=file_name,
                                    key=f"download_{file_name}"
                                )

                            except Exception as e:

                                st.error(
                                    f"❌ Download failed: {e}"
                                )

                    with col2:

                        if st.button(
                            "🗑️ Delete",
                            key=f"delete_{file_name}"
                        ):

                            try:

                                (
                                    supabase
                                    .storage
                                    .from_(
                                        STORAGE_BUCKET
                                    )
                                    .remove(
                                        [file_name]
                                    )
                                )

                                st.success(
                                    "✅ File deleted."
                                )

                                st.rerun()

                            except Exception as e:

                                st.error(
                                    f"❌ Delete failed: {e}"
                                )

                    st.markdown("---")

            else:

                st.info(
                    "📭 No uploaded certificates found."
                )

        except Exception as e:

            st.error(
                f"❌ Unable to access storage: {e}"
            )


# ============================================================
# FOOTER
# ============================================================

st.markdown(
    """
    <div class="footer">
        🎓 Student Certificate Management System
        <br>
        Built with Python, Streamlit & Supabase
    </div>
    """,
    unsafe_allow_html=True
)
